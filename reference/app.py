#!/usr/bin/env python3
"""
Dimak Prospector -- local web app
=================================

A point-and-click prospecting tool for Dimak Kapi's Gulf fire door export.
You run it once on your own machine; then you use it entirely in the browser:
search for companies, review them, mark who you've contacted, and export.

Runs locally so your Google API key stays on your machine (never exposed to a
browser) and so the app can visit company websites for emails (browsers can't).

SETUP  (once)
    pip install flask requests beautifulsoup4 openpyxl
    export GOOGLE_PLACES_API_KEY="your-key"     # Windows: set GOOGLE_PLACES_API_KEY=your-key

START
    python app.py
    -> open the printed address (http://127.0.0.1:5000) in your browser

Your data lives in prospects.db next to this file. Delete it to start fresh.
Run  python app.py --self-test  to check the logic offline (no key needed).
"""

import os
import re
import sys
import time
import sqlite3
from datetime import date

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prospects.db")

CITIES = [
    {"city": "Dubai",       "country": "UAE",          "region": "AE"},
    {"city": "Abu Dhabi",   "country": "UAE",          "region": "AE"},
    {"city": "Sharjah",     "country": "UAE",          "region": "AE"},
    {"city": "Riyadh",      "country": "Saudi Arabia", "region": "SA"},
    {"city": "Jeddah",      "country": "Saudi Arabia", "region": "SA"},
    {"city": "Dammam",      "country": "Saudi Arabia", "region": "SA"},
    {"city": "Doha",        "country": "Qatar",        "region": "QA"},
    {"city": "Kuwait City", "country": "Kuwait",       "region": "KW"},
    {"city": "Muscat",      "country": "Oman",         "region": "OM"},
    {"city": "Manama",      "country": "Bahrain",      "region": "BH"},
]

SEGMENTS = [
    "Distributor / Trading",
    "Contractor (general / fit-out)",
    "Architect / Specifier",
    "Facility / FM",
]

TERM_SUGGESTIONS = [
    "fire door supplier", "fire rated door distributor", "steel door supplier",
    "doors and hardware supplier", "building materials trading company",
    "architectural hardware supplier", "fit out contractor",
    "interior fit out company", "joinery contractor", "general contracting company",
    "architecture firm", "architectural consultant", "facilities management company",
]

STATUSES = ["New", "Contacted", "Replied", "Not a fit"]

PLACES_URL = "https://places.googleapis.com/v1/places:searchText"
FIELD_MASK = ",".join([
    "places.id", "places.displayName", "places.formattedAddress",
    "places.nationalPhoneNumber", "places.internationalPhoneNumber",
    "places.websiteUri", "places.rating", "places.userRatingCount",
    "places.primaryTypeDisplayName", "places.googleMapsUri", "nextPageToken",
])
MAX_PAGES_PER_QUERY = 3

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
EMAIL_BLOCKLIST = ("example.com", "sentry.io", "wixpress.com", "@2x", ".png", ".jpg")

COLUMNS = ["place_id", "company", "segment", "country", "city", "category",
           "address", "phone", "website", "emails", "rating", "reviews",
           "google_maps_url", "status", "notes", "source", "collected"]


# ----------------------------------------------------------------------------
# DATABASE
# ----------------------------------------------------------------------------

def db_connect(path=DB_PATH):
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    return conn


def db_init(path=DB_PATH):
    conn = db_connect(path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prospects (
            place_id TEXT PRIMARY KEY, company TEXT, segment TEXT, country TEXT,
            city TEXT, category TEXT, address TEXT, phone TEXT, website TEXT,
            emails TEXT, rating REAL, reviews INTEGER, google_maps_url TEXT,
            status TEXT DEFAULT 'New', notes TEXT DEFAULT '', source TEXT,
            collected TEXT
        )""")
    conn.commit()
    return conn


def upsert(conn, row):
    """Insert a prospect, or refresh contact fields + merge segment tags if it
    already exists. Never clobbers the user's status or notes."""
    cur = conn.execute("SELECT segment FROM prospects WHERE place_id = ?",
                       (row["place_id"],))
    existing = cur.fetchone()
    if existing:
        tags = set((existing["segment"] or "").split(" | ")) if existing["segment"] else set()
        tags.add(row["segment"])
        conn.execute("""UPDATE prospects SET segment=?, phone=?, website=?,
                        rating=?, reviews=? WHERE place_id=?""",
                     (" | ".join(sorted(t for t in tags if t)), row["phone"],
                      row["website"], row["rating"], row["reviews"], row["place_id"]))
        return "updated"
    conn.execute(f"""INSERT INTO prospects ({",".join(COLUMNS)})
                     VALUES ({",".join("?" * len(COLUMNS))})""",
                 [row.get(c, "") for c in COLUMNS])
    return "inserted"


# ----------------------------------------------------------------------------
# PURE LOGIC  (offline-testable)
# ----------------------------------------------------------------------------

def extract_emails(html):
    if not html:
        return []
    out = []
    for m in EMAIL_RE.findall(html):
        e = m.lower().rstrip(".")
        if any(bad in e for bad in EMAIL_BLOCKLIST):
            continue
        if e not in out:
            out.append(e)
    return out


def wa_link(phone):
    digits = re.sub(r"\D", "", phone or "")
    return f"https://wa.me/{digits}" if digits else ""


# ----------------------------------------------------------------------------
# NETWORK  (lazy imports)
# ----------------------------------------------------------------------------

def places_search(query, region, api_key):
    import requests
    rows, token, pages = [], None, 0
    while pages < MAX_PAGES_PER_QUERY:
        body = {"textQuery": query, "regionCode": region, "languageCode": "en"}
        if token:
            body["pageToken"] = token
        r = requests.post(PLACES_URL, headers={
            "Content-Type": "application/json", "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": FIELD_MASK}, json=body, timeout=30)
        if r.status_code != 200:
            raise RuntimeError(f"Places API {r.status_code}: {r.text[:200]}")
        data = r.json()
        for p in data.get("places", []):
            rows.append({
                "place_id": p.get("id", ""),
                "company": (p.get("displayName") or {}).get("text", ""),
                "category": (p.get("primaryTypeDisplayName") or {}).get("text", ""),
                "address": p.get("formattedAddress", ""),
                "phone": p.get("internationalPhoneNumber") or p.get("nationalPhoneNumber", ""),
                "website": p.get("websiteUri", ""),
                "rating": p.get("rating", ""),
                "reviews": p.get("userRatingCount", ""),
                "google_maps_url": p.get("googleMapsUri", ""),
            })
        token = data.get("nextPageToken")
        pages += 1
        if not token:
            break
        time.sleep(2)
    return rows


def enrich_email(website):
    import requests
    if not website:
        return ""
    base = website.rstrip("/")
    emails = []
    ua = {"User-Agent": "Mozilla/5.0 (compatible; DimakProspector/1.0)"}
    for path in ("", "/contact", "/contact-us", "/iletisim", "/about"):
        try:
            r = requests.get(base + path, headers=ua, timeout=12)
            if r.status_code == 200:
                for e in extract_emails(r.text):
                    if e not in emails:
                        emails.append(e)
            if emails:
                break
        except Exception:
            continue
        time.sleep(0.8)
    return " | ".join(emails[:3])


# ----------------------------------------------------------------------------
# FLASK APP
# ----------------------------------------------------------------------------

def create_app():
    from flask import Flask, request, jsonify, Response
    app = Flask(__name__)
    db_init()

    @app.route("/")
    def index():
        return Response(PAGE, mimetype="text/html")

    @app.route("/api/config")
    def config():
        return jsonify(cities=CITIES, segments=SEGMENTS,
                       terms=TERM_SUGGESTIONS, statuses=STATUSES)

    @app.route("/api/search", methods=["POST"])
    def search():
        api_key = os.environ.get("GOOGLE_PLACES_API_KEY")
        if not api_key:
            return jsonify(error="GOOGLE_PLACES_API_KEY is not set on the machine "
                                 "running this app."), 400
        body = request.get_json(force=True)
        keyword = (body.get("keyword") or "").strip()
        segment = body.get("segment") or "Unclassified"
        want = body.get("cities") or [c["city"] for c in CITIES]
        enrich = bool(body.get("enrich"))
        if not keyword:
            return jsonify(error="Type what to search for first."), 400

        chosen = [c for c in CITIES if c["city"] in want]
        conn = db_init()
        added = updated = 0
        try:
            for loc in chosen:
                hits = places_search(f"{keyword} in {loc['city']}", loc["region"], api_key)
                for h in hits:
                    if not h["place_id"]:
                        continue
                    h.update(segment=segment, country=loc["country"], city=loc["city"],
                             emails="", source="Google Places",
                             collected=date.today().isoformat())
                    if enrich and h["website"]:
                        h["emails"] = enrich_email(h["website"])
                    result = upsert(conn, h)
                    added += result == "inserted"
                    updated += result == "updated"
            conn.commit()
        except RuntimeError as e:
            return jsonify(error=str(e)), 502
        return jsonify(added=added, updated=updated)

    @app.route("/api/prospects")
    def prospects():
        conn = db_init()
        clauses, params = [], []
        for field in ("country", "segment", "status"):
            val = request.args.get(field)
            if val and val != "All":
                if field == "segment":
                    clauses.append("segment LIKE ?"); params.append(f"%{val}%")
                else:
                    clauses.append(f"{field} = ?"); params.append(val)
        q = request.args.get("q")
        if q:
            clauses.append("(company LIKE ? OR city LIKE ?)")
            params += [f"%{q}%", f"%{q}%"]
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        rows = conn.execute(
            f"SELECT * FROM prospects {where} ORDER BY country, city, company",
            params).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            d["wa"] = wa_link(d["phone"])
            out.append(d)
        counts = {s: 0 for s in STATUSES}
        for d in out:
            counts[d["status"]] = counts.get(d["status"], 0) + 1
        return jsonify(rows=out, total=len(out), counts=counts)

    @app.route("/api/update", methods=["POST"])
    def update():
        body = request.get_json(force=True)
        pid = body.get("place_id")
        conn = db_init()
        if "status" in body:
            conn.execute("UPDATE prospects SET status=? WHERE place_id=?",
                         (body["status"], pid))
        if "notes" in body:
            conn.execute("UPDATE prospects SET notes=? WHERE place_id=?",
                         (body["notes"], pid))
        conn.commit()
        return jsonify(ok=True)

    @app.route("/api/enrich", methods=["POST"])
    def enrich_one():
        pid = request.get_json(force=True).get("place_id")
        conn = db_init()
        row = conn.execute("SELECT website FROM prospects WHERE place_id=?",
                          (pid,)).fetchone()
        if not row or not row["website"]:
            return jsonify(emails="")
        emails = enrich_email(row["website"])
        conn.execute("UPDATE prospects SET emails=? WHERE place_id=?", (emails, pid))
        conn.commit()
        return jsonify(emails=emails)

    @app.route("/api/export")
    def export():
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        conn = db_init()
        rows = conn.execute("SELECT * FROM prospects ORDER BY country, city, company").fetchall()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Prospects"
        headers = ["Company", "Segment", "Country", "City", "Category", "Address",
                   "Phone", "Website", "Emails", "Rating", "Reviews", "Status",
                   "Notes", "Maps URL"]
        keys = ["company", "segment", "country", "city", "category", "address",
                "phone", "website", "emails", "rating", "reviews", "status",
                "notes", "google_maps_url"]
        fill = PatternFill("solid", fgColor="C8511B"); f = Font(bold=True, color="FFFFFF")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.fill = fill; cell.font = f
        for i, r in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                ws.cell(i, c, r[k])
        for c, w in enumerate([32, 26, 14, 13, 20, 42, 20, 32, 32, 8, 9, 12, 30, 30], 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        path = os.path.join(os.path.dirname(DB_PATH),
                            f"dimak_prospects_{date.today().isoformat()}.xlsx")
        wb.save(path)
        with open(path, "rb") as fh:
            data = fh.read()
        return Response(data, mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            headers={"Content-Disposition":
                     f"attachment; filename={os.path.basename(path)}"})

    return app


# ----------------------------------------------------------------------------
# FRONTEND  (single self-contained page, no external assets)
# ----------------------------------------------------------------------------

PAGE = r"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dimak Prospector</title>
<style>
:root{
  --bg:#eef1f4; --panel:#ffffff; --ink:#161a20; --steel:#5b636e; --mute:#8b939f;
  --line:#dde2e8; --ember:#d2541c; --ember-dk:#9c3d12;
  --new:#7b8592; --contacted:#c9820e; --replied:#1f8a54; --nofit:#b04a3f;
}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--ink);
  font-family:system-ui,-apple-system,Segoe UI,Roboto,sans-serif;font-size:14px}
.mono{font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace}
header{background:var(--ink);color:#fff;padding:14px 22px;display:flex;
  align-items:baseline;gap:14px;border-bottom:3px solid var(--ember)}
header h1{margin:0;font-size:17px;letter-spacing:.14em;text-transform:uppercase;font-weight:700}
header .sub{color:#9aa3af;font-size:12px;letter-spacing:.03em}
.wrap{max-width:1220px;margin:0 auto;padding:20px 22px 60px}
.panel{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:16px 18px;margin-bottom:16px}
.panel h2{margin:0 0 12px;font-size:11px;letter-spacing:.16em;text-transform:uppercase;color:var(--mute)}
.row{display:flex;flex-wrap:wrap;gap:10px;align-items:flex-end}
label{display:block;font-size:11px;letter-spacing:.05em;color:var(--steel);margin-bottom:4px}
input[type=text],select{padding:9px 10px;border:1px solid var(--line);border-radius:6px;
  font-size:14px;background:#fff;color:var(--ink);min-width:150px}
input[type=text]:focus,select:focus{outline:2px solid var(--ember);outline-offset:-1px;border-color:var(--ember)}
.grow{flex:1;min-width:220px}
.cities{display:flex;flex-wrap:wrap;gap:6px;margin-top:4px}
.chip{border:1px solid var(--line);border-radius:999px;padding:5px 11px;cursor:pointer;
  user-select:none;font-size:12px;background:#fff;color:var(--steel)}
.chip.on{background:var(--ink);color:#fff;border-color:var(--ink)}
.btn{background:var(--ember);color:#fff;border:none;border-radius:6px;padding:10px 18px;
  font-weight:600;cursor:pointer;font-size:14px}
.btn:hover{background:var(--ember-dk)} .btn:disabled{opacity:.5;cursor:default}
.btn.ghost{background:#fff;color:var(--ink);border:1px solid var(--line)}
.check{display:flex;align-items:center;gap:6px;font-size:13px;color:var(--steel)}
.stats{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px}
.stat{background:var(--panel);border:1px solid var(--line);border-radius:6px;padding:8px 14px;min-width:92px}
.stat b{display:block;font-size:20px} .stat span{font-size:11px;color:var(--mute);letter-spacing:.05em}
.filters{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;margin-bottom:12px}
table{width:100%;border-collapse:collapse;background:var(--panel);border:1px solid var(--line);border-radius:8px;overflow:hidden}
th{background:#f6f8fa;text-align:left;font-size:10px;letter-spacing:.1em;text-transform:uppercase;
  color:var(--mute);padding:10px;border-bottom:1px solid var(--line)}
td{padding:10px;border-bottom:1px solid var(--line);vertical-align:top}
tr:last-child td{border-bottom:none}
td .co{font-weight:600} td .meta{color:var(--mute);font-size:12px}
a{color:var(--ember-dk);text-decoration:none} a:hover{text-decoration:underline}
.pill{font-size:11px;padding:3px 9px;border-radius:999px;border:1px solid transparent;font-weight:600}
select.status{border-radius:999px;padding:4px 8px;font-size:12px;font-weight:600;border:1px solid var(--line)}
.s-New{color:var(--new)} .s-Contacted{color:var(--contacted)}
.s-Replied{color:var(--replied)} .s-Notafit{color:var(--nofit)}
.notes{width:100%;border:1px solid var(--line);border-radius:5px;padding:5px;font-size:12px;font-family:inherit;resize:vertical;min-height:34px}
.empty{padding:40px;text-align:center;color:var(--mute)}
.spin{display:inline-block;width:13px;height:13px;border:2px solid #fff;border-top-color:transparent;border-radius:50%;animation:sp .7s linear infinite;vertical-align:-2px;margin-right:6px}
@keyframes sp{to{transform:rotate(360deg)}}
.note{font-size:12px;color:var(--mute)}
</style></head><body>
<header><h1>Dimak Prospector</h1><span class="sub">Gulf fire door lead pipeline</span></header>
<div class="wrap">

  <div class="panel">
    <h2>Find companies</h2>
    <div class="row">
      <div class="grow">
        <label>What to search for</label>
        <input type="text" id="keyword" list="terms" placeholder="e.g. fire door supplier">
        <datalist id="terms"></datalist>
      </div>
      <div>
        <label>Tag results as</label>
        <select id="segment"></select>
      </div>
      <label class="check"><input type="checkbox" id="enrich"> Look up emails (slower)</label>
      <button class="btn" id="go">Search</button>
    </div>
    <div style="margin-top:12px">
      <label>Cities <a href="#" id="allCities" class="note">all</a> / <a href="#" id="noCities" class="note">none</a></label>
      <div class="cities" id="cities"></div>
    </div>
    <div id="searchMsg" class="note" style="margin-top:10px"></div>
  </div>

  <div class="stats" id="stats"></div>

  <div class="filters">
    <div><label>Country</label><select id="fCountry"></select></div>
    <div><label>Segment</label><select id="fSegment"></select></div>
    <div><label>Status</label><select id="fStatus"></select></div>
    <div class="grow"><label>Find in list</label><input type="text" id="fq" placeholder="company or city"></div>
    <button class="btn ghost" id="exportBtn">Export to Excel</button>
  </div>

  <div id="tableWrap"></div>
</div>

<script>
let CFG={};
const $=s=>document.querySelector(s);

async function boot(){
  CFG=await (await fetch('/api/config')).json();
  $('#segment').innerHTML=CFG.segments.map(s=>`<option>${s}</option>`).join('');
  $('#terms').innerHTML=CFG.terms.map(t=>`<option value="${t}">`).join('');
  $('#cities').innerHTML=CFG.cities.map(c=>
    `<span class="chip on" data-city="${c.city}">${c.city}</span>`).join('');
  const countries=['All',...new Set(CFG.cities.map(c=>c.country))];
  $('#fCountry').innerHTML=countries.map(c=>`<option>${c}</option>`).join('');
  $('#fSegment').innerHTML=['All',...CFG.segments].map(s=>`<option>${s}</option>`).join('');
  $('#fStatus').innerHTML=['All',...CFG.statuses].map(s=>`<option>${s}</option>`).join('');
  document.querySelectorAll('.chip').forEach(ch=>ch.onclick=()=>ch.classList.toggle('on'));
  $('#allCities').onclick=e=>{e.preventDefault();document.querySelectorAll('.chip').forEach(c=>c.classList.add('on'))};
  $('#noCities').onclick=e=>{e.preventDefault();document.querySelectorAll('.chip').forEach(c=>c.classList.remove('on'))};
  $('#go').onclick=runSearch;
  ['#fCountry','#fSegment','#fStatus'].forEach(s=>$(s).onchange=load);
  $('#fq').oninput=()=>{clearTimeout(window._t);window._t=setTimeout(load,300)};
  $('#exportBtn').onclick=()=>location.href='/api/export';
  load();
}

async function runSearch(){
  const keyword=$('#keyword').value.trim();
  if(!keyword){$('#searchMsg').textContent='Type what to search for first.';return;}
  const cities=[...document.querySelectorAll('.chip.on')].map(c=>c.dataset.city);
  if(!cities.length){$('#searchMsg').textContent='Pick at least one city.';return;}
  const btn=$('#go');btn.disabled=true;btn.innerHTML='<span class="spin"></span>Searching';
  $('#searchMsg').textContent='';
  try{
    const res=await fetch('/api/search',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({keyword,segment:$('#segment').value,cities,enrich:$('#enrich').checked})});
    const d=await res.json();
    if(!res.ok){$('#searchMsg').textContent=d.error||'Search failed.';}
    else{$('#searchMsg').textContent=`Added ${d.added} new, refreshed ${d.updated}.`;await load();}
  }catch(e){$('#searchMsg').textContent='Could not reach the app. Is it still running?';}
  btn.disabled=false;btn.textContent='Search';
}

async function load(){
  const p=new URLSearchParams({country:$('#fCountry').value,segment:$('#fSegment').value,
    status:$('#fStatus').value,q:$('#fq').value});
  const d=await (await fetch('/api/prospects?'+p)).json();
  $('#stats').innerHTML=`<div class="stat"><b>${d.total}</b><span>SHOWING</span></div>`+
    CFG.statuses.map(s=>`<div class="stat"><b>${d.counts[s]||0}</b><span>${s.toUpperCase()}</span></div>`).join('');
  if(!d.rows.length){$('#tableWrap').innerHTML='<div class="panel empty">No prospects yet. Run a search above.</div>';return;}
  const body=d.rows.map(r=>{
    const wa=r.wa?`<a href="${r.wa}" target="_blank">WhatsApp</a>`:'';
    const site=r.website?`<a href="${r.website}" target="_blank">site</a>`:'';
    const mail=r.emails?r.emails.split(' | ').map(e=>`<a href="mailto:${e}">${e}</a>`).join('<br>'):'<span class="meta">—</span>';
    const opts=CFG.statuses.map(s=>`<option ${s===r.status?'selected':''}>${s}</option>`).join('');
    return `<tr>
      <td><div class="co">${r.company||''}</div><div class="meta">${r.category||''}</div>
          <a href="${r.google_maps_url}" target="_blank" class="meta">map</a></td>
      <td class="meta">${r.segment||''}</td>
      <td>${r.country||''}<div class="meta">${r.city||''}</div></td>
      <td class="mono">${r.phone||''}<div>${wa} ${site}</div></td>
      <td class="mono" style="font-size:12px">${mail}</td>
      <td class="mono">${r.rating||''}<div class="meta">${r.reviews||0}</div></td>
      <td><select class="status s-${(r.status||'').replace(/ /g,'')}" data-id="${r.place_id}">${opts}</select></td>
      <td style="min-width:150px"><textarea class="notes" data-id="${r.place_id}">${r.notes||''}</textarea></td>
    </tr>`;}).join('');
  $('#tableWrap').innerHTML=`<table><thead><tr>
    <th>Company</th><th>Segment</th><th>Location</th><th>Phone</th><th>Email</th>
    <th>Rating</th><th>Status</th><th>Notes</th></tr></thead><tbody>${body}</tbody></table>`;
  document.querySelectorAll('select.status').forEach(s=>s.onchange=()=>{
    save(s.dataset.id,{status:s.value});
    s.className='status s-'+s.value.replace(/ /g,'');
  });
  document.querySelectorAll('textarea.notes').forEach(t=>t.onblur=()=>save(t.dataset.id,{notes:t.value}));
}

async function save(place_id,fields){
  await fetch('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({place_id,...fields})});
}
boot();
</script></body></html>"""


# ----------------------------------------------------------------------------
# OFFLINE SELF-TEST
# ----------------------------------------------------------------------------

def self_test():
    import tempfile
    tmp = os.path.join(tempfile.mkdtemp(), "t.db")
    conn = db_init(tmp)
    base = {"company": "Gulf Doors LLC", "country": "UAE", "city": "Dubai",
            "category": "Supplier", "address": "x", "phone": "+971 4 123 4567",
            "website": "http://g.ae", "rating": 4.5, "reviews": 12,
            "google_maps_url": "http://m", "emails": "", "source": "Google Places",
            "collected": "2026-07-14"}
    r1 = dict(base, place_id="A", segment="Distributor / Trading")
    assert upsert(conn, r1) == "inserted"
    r2 = dict(base, place_id="A", segment="Contractor (general / fit-out)")
    assert upsert(conn, r2) == "updated"
    conn.commit()
    seg = conn.execute("SELECT segment FROM prospects WHERE place_id='A'").fetchone()["segment"]
    assert seg == "Contractor (general / fit-out) | Distributor / Trading", seg

    conn.execute("UPDATE prospects SET status='Contacted', notes='called' WHERE place_id='A'")
    conn.commit()
    row = conn.execute("SELECT status, notes FROM prospects WHERE place_id='A'").fetchone()
    assert row["status"] == "Contacted" and row["notes"] == "called"

    # re-finding must NOT wipe status/notes
    upsert(conn, dict(base, place_id="A", segment="Architect / Specifier"))
    conn.commit()
    row = conn.execute("SELECT status, notes FROM prospects WHERE place_id='A'").fetchone()
    assert row["status"] == "Contacted" and row["notes"] == "called"

    assert wa_link("+971 4 123 4567") == "https://wa.me/97141234567"
    emails = extract_emails('a <a href="mailto:sales@g.ae">x</a> noreply@example.com logo@2x.png')
    assert emails == ["sales@g.ae"], emails

    print("SELF-TEST PASSED")
    print(f"  segment tags merged: {seg}")
    print("  status/notes survive re-discovery: yes")
    print(f"  whatsapp link + email parse: ok ({emails})")


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        self_test()
    else:
        if not os.environ.get("GOOGLE_PLACES_API_KEY"):
            print("Note: GOOGLE_PLACES_API_KEY is not set. The app will start, "
                  "but searches will fail until you set it.")
        print("Dimak Prospector running -> open http://127.0.0.1:5000")
        create_app().run(host="127.0.0.1", port=5000, debug=False)
