#!/usr/bin/env python3
"""
Gulf Fire Door Prospector
=========================

Finds B2B prospects for Dimak Kapi (fire doors) across the Gulf, live from
Google Places, and exports a review-ready Excel workbook.

WHAT IT DOES
    1. Runs a grid of (search term x city) queries against the Google Places API.
    2. Collects each business: name, address, phone, website, rating, category.
    3. Tags each row with segment + city + country.
    4. Optionally visits each website to pull public contact emails.
    5. De-duplicates and writes an Excel workbook with filters + a summary sheet.

WHY PLACES API (not a scraper): it is the freshest source (Google's own live
data), legal, structured, and needs no proxy setup. Use Bright Data for the
sources Places cannot reach (tenders, trade-show exhibitor lists, directories).

SETUP
    pip install requests beautifulsoup4 openpyxl
    export GOOGLE_PLACES_API_KEY="your-key"
        (Google Cloud Console -> enable "Places API (New)" -> create API key)

RUN
    python gulf_fire_door_prospector.py                 # harvest only
    python gulf_fire_door_prospector.py --enrich-emails # + visit sites for emails
    python gulf_fire_door_prospector.py --self-test     # offline logic check

COST NOTE
    The contact-detail fields (website, phone) are billed at a higher Places
    tier than name/address. Trim FIELD_MASK if you want to lower per-request
    cost. Google gives a recurring monthly free credit; verify current pricing
    before running the full grid.
"""

import os
import re
import sys
import time
import argparse
from datetime import date

# ----------------------------------------------------------------------------
# CONFIG  -- edit these lists freely; they are the whole strategy of the tool
# ----------------------------------------------------------------------------

# Search terms grouped by prospect segment. The segment label is carried onto
# every result so you can filter your outreach by who you are talking to.
SEARCH_TERMS = {
    "Distributor / Trading": [
        "fire door supplier",
        "fire rated door distributor",
        "steel door supplier",
        "doors and hardware supplier",
        "building materials trading company",
        "architectural hardware supplier",
    ],
    "Contractor (general / fit-out)": [
        "fit out contractor",
        "interior fit out company",
        "joinery contractor",
        "general contracting company",
        "MEP and civil contractor",
    ],
    "Architect / Specifier": [
        "architecture firm",
        "architectural consultant",
        "interior design consultancy",
    ],
    "Facility / FM": [
        "facilities management company",
        "hotel group head office",
        "hospital facilities management",
    ],
}

# Gulf cities with their ISO region codes (biases Places results to that market).
CITIES = [
    {"city": "Dubai",        "country": "UAE",     "region": "AE"},
    {"city": "Abu Dhabi",    "country": "UAE",     "region": "AE"},
    {"city": "Sharjah",      "country": "UAE",     "region": "AE"},
    {"city": "Riyadh",       "country": "Saudi Arabia", "region": "SA"},
    {"city": "Jeddah",       "country": "Saudi Arabia", "region": "SA"},
    {"city": "Dammam",       "country": "Saudi Arabia", "region": "SA"},
    {"city": "Doha",         "country": "Qatar",   "region": "QA"},
    {"city": "Kuwait City",  "country": "Kuwait",  "region": "KW"},
    {"city": "Muscat",       "country": "Oman",    "region": "OM"},
    {"city": "Manama",       "country": "Bahrain", "region": "BH"},
]

MAX_PAGES_PER_QUERY = 3        # Places returns up to 20/page, 3 pages = 60 max
REQUEST_PAUSE_SEC   = 0.4      # politeness between API calls
ENRICH_PAUSE_SEC    = 1.0      # politeness between website visits
OUTPUT_FILE = f"gulf_fire_door_prospects_{date.today().isoformat()}.xlsx"

PLACES_URL = "https://places.googleapis.com/v1/places:searchText"
FIELD_MASK = ",".join([
    "places.id",
    "places.displayName",
    "places.formattedAddress",
    "places.nationalPhoneNumber",
    "places.internationalPhoneNumber",
    "places.websiteUri",
    "places.rating",
    "places.userRatingCount",
    "places.primaryTypeDisplayName",
    "places.googleMapsUri",
    "nextPageToken",
])

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# Junk addresses that show up in page source but are never real contacts.
EMAIL_BLOCKLIST = ("example.com", "sentry.io", "wixpress.com", "@2x", ".png", ".jpg")


# ----------------------------------------------------------------------------
# PURE LOGIC  (no network -- covered by --self-test)
# ----------------------------------------------------------------------------

def build_query_plan():
    """Expand the term x city grid into a flat list of query jobs."""
    plan = []
    for segment, terms in SEARCH_TERMS.items():
        for term in terms:
            for loc in CITIES:
                plan.append({
                    "segment": segment,
                    "query": f"{term} in {loc['city']}",
                    "city": loc["city"],
                    "country": loc["country"],
                    "region": loc["region"],
                })
    return plan


def dedupe(rows):
    """Keep one row per Google place id; merge segment tags if a company shows
    up under more than one search so you keep the full picture of who they are."""
    by_id = {}
    for r in rows:
        pid = r.get("place_id")
        if not pid:
            continue
        if pid in by_id:
            existing = by_id[pid]
            tags = set(existing["segment"].split(" | "))
            tags.add(r["segment"])
            existing["segment"] = " | ".join(sorted(tags))
        else:
            by_id[pid] = dict(r)
    return list(by_id.values())


def extract_emails(html):
    """Pull plausible public contact emails from page HTML/text."""
    if not html:
        return []
    found = []
    for match in EMAIL_RE.findall(html):
        m = match.lower().rstrip(".")
        if any(bad in m for bad in EMAIL_BLOCKLIST):
            continue
        if m not in found:
            found.append(m)
    return found


# ----------------------------------------------------------------------------
# NETWORK  (needs your key / internet -- lazy imports so --self-test stays clean)
# ----------------------------------------------------------------------------

def places_search(job, api_key):
    """Run one Places text search (with pagination) and return normalized rows."""
    import requests

    rows, page_token, pages = [], None, 0
    while pages < MAX_PAGES_PER_QUERY:
        body = {"textQuery": job["query"], "regionCode": job["region"],
                "languageCode": "en"}
        if page_token:
            body["pageToken"] = page_token
        resp = requests.post(
            PLACES_URL,
            headers={"Content-Type": "application/json",
                     "X-Goog-Api-Key": api_key,
                     "X-Goog-FieldMask": FIELD_MASK},
            json=body, timeout=30,
        )
        if resp.status_code != 200:
            print(f"    ! {resp.status_code} for '{job['query']}': {resp.text[:160]}")
            break
        data = resp.json()
        for p in data.get("places", []):
            rows.append({
                "company": (p.get("displayName") or {}).get("text", ""),
                "segment": job["segment"],
                "country": job["country"],
                "city": job["city"],
                "category": (p.get("primaryTypeDisplayName") or {}).get("text", ""),
                "address": p.get("formattedAddress", ""),
                "phone": p.get("internationalPhoneNumber")
                         or p.get("nationalPhoneNumber", ""),
                "website": p.get("websiteUri", ""),
                "emails": "",
                "rating": p.get("rating", ""),
                "reviews": p.get("userRatingCount", ""),
                "google_maps_url": p.get("googleMapsUri", ""),
                "place_id": p.get("id", ""),
                "source": "Google Places",
                "collected": date.today().isoformat(),
            })
        page_token = data.get("nextPageToken")
        pages += 1
        if not page_token:
            break
        time.sleep(2)  # token needs a moment to become valid
    return rows


def enrich_email(website):
    """Visit a company site (home + common contact paths) and pull emails."""
    import requests
    if not website:
        return ""
    base = website.rstrip("/")
    paths = ["", "/contact", "/contact-us", "/contactus", "/iletisim", "/about"]
    emails = []
    headers = {"User-Agent": "Mozilla/5.0 (compatible; DimakProspector/1.0)"}
    for path in paths:
        try:
            r = requests.get(base + path, headers=headers, timeout=12)
            if r.status_code == 200:
                for e in extract_emails(r.text):
                    if e not in emails:
                        emails.append(e)
            if emails:
                break  # got something on the first productive page
        except requests.RequestException:
            continue
        time.sleep(ENRICH_PAUSE_SEC)
    return " | ".join(emails[:3])


def harvest(api_key, enrich=False):
    plan = build_query_plan()
    print(f"Planned {len(plan)} queries across {len(CITIES)} cities.")
    raw = []
    for i, job in enumerate(plan, 1):
        print(f"[{i}/{len(plan)}] {job['query']}")
        raw.extend(places_search(job, api_key))
        time.sleep(REQUEST_PAUSE_SEC)

    rows = dedupe(raw)
    print(f"\n{len(raw)} raw hits -> {len(rows)} unique companies.")

    if enrich:
        with_site = [r for r in rows if r["website"]]
        print(f"Enriching emails for {len(with_site)} sites with a website...")
        for j, r in enumerate(with_site, 1):
            r["emails"] = enrich_email(r["website"])
            if j % 25 == 0:
                print(f"    ...{j}/{len(with_site)}")
    return rows


# ----------------------------------------------------------------------------
# OUTPUT
# ----------------------------------------------------------------------------

def write_excel(rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cols = ["company", "segment", "country", "city", "category", "address",
            "phone", "website", "emails", "rating", "reviews",
            "google_maps_url", "place_id", "source", "collected"]
    headers = ["Company", "Segment", "Country", "City", "Category", "Address",
               "Phone", "Website", "Emails", "Rating", "Reviews",
               "Maps URL", "Place ID", "Source", "Collected"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"
    head_fill = PatternFill("solid", fgColor="C8511B")   # Dimak orange
    head_font = Font(bold=True, color="FFFFFF")
    for c, label in enumerate(headers, 1):
        cell = ws.cell(1, c, label)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    for r, row in enumerate(sorted(rows, key=lambda x: (x["country"], x["city"], x["company"])), 2):
        for c, key in enumerate(cols, 1):
            ws.cell(r, c, row.get(key, ""))
    widths = [34, 26, 14, 14, 22, 44, 20, 34, 34, 8, 9, 30, 30, 14, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"

    # Summary sheet: counts by country x segment
    summary = wb.create_sheet("Summary")
    counts = {}
    for row in rows:
        for seg in row["segment"].split(" | "):
            counts[(row["country"], seg)] = counts.get((row["country"], seg), 0) + 1
    summary.cell(1, 1, "Country").font = Font(bold=True)
    summary.cell(1, 2, "Segment").font = Font(bold=True)
    summary.cell(1, 3, "Companies").font = Font(bold=True)
    for i, ((country, seg), n) in enumerate(sorted(counts.items()), 2):
        summary.cell(i, 1, country)
        summary.cell(i, 2, seg)
        summary.cell(i, 3, n)
    for col, w in zip("ABC", (16, 30, 12)):
        summary.column_dimensions[col].width = w

    wb.save(path)
    print(f"Saved {len(rows)} prospects -> {path}")


# ----------------------------------------------------------------------------
# SELF-TEST  (offline: proves the logic without a key or internet)
# ----------------------------------------------------------------------------

def self_test():
    plan = build_query_plan()
    expected = sum(len(t) for t in SEARCH_TERMS.values()) * len(CITIES)
    assert len(plan) == expected, (len(plan), expected)
    assert plan[0]["segment"] == "Distributor / Trading"
    assert " in " in plan[0]["query"]

    sample = [
        {"place_id": "A", "company": "Gulf Doors LLC", "segment": "Distributor / Trading",
         "country": "UAE", "city": "Dubai", "website": "x"},
        {"place_id": "A", "company": "Gulf Doors LLC", "segment": "Contractor (general / fit-out)",
         "country": "UAE", "city": "Dubai", "website": "x"},
        {"place_id": "B", "company": "Doha Fitout", "segment": "Contractor (general / fit-out)",
         "country": "Qatar", "city": "Doha", "website": ""},
    ]
    merged = dedupe(sample)
    assert len(merged) == 2
    a = next(r for r in merged if r["place_id"] == "A")
    assert a["segment"] == "Contractor (general / fit-out) | Distributor / Trading"

    html = ('contact <a href="mailto:sales@gulfdoors.ae">us</a> or '
            'info@gulfdoors.ae. ignore noreply@example.com and logo@2x.png')
    emails = extract_emails(html)
    assert "sales@gulfdoors.ae" in emails and "info@gulfdoors.ae" in emails
    assert not any("example.com" in e or "2x" in e for e in emails)

    print("SELF-TEST PASSED")
    print(f"  query plan: {len(plan)} jobs")
    print(f"  dedupe merged 3 rows -> {len(merged)} companies, tags combined")
    print(f"  emails extracted: {emails}")


# ----------------------------------------------------------------------------
# ENTRY
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Gulf fire door prospector")
    ap.add_argument("--enrich-emails", action="store_true",
                    help="visit each website to pull public contact emails")
    ap.add_argument("--self-test", action="store_true",
                    help="run offline logic checks and exit")
    ap.add_argument("--out", default=OUTPUT_FILE, help="output .xlsx path")
    args = ap.parse_args()

    if args.self_test:
        self_test()
        return

    api_key = os.environ.get("GOOGLE_PLACES_API_KEY")
    if not api_key:
        sys.exit("Set GOOGLE_PLACES_API_KEY first "
                 "(export GOOGLE_PLACES_API_KEY=...).")

    rows = harvest(api_key, enrich=args.enrich_emails)
    write_excel(rows, args.out)


if __name__ == "__main__":
    main()
